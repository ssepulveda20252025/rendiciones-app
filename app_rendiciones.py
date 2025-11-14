import streamlit as st
import pandas as pd
import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# CONFIGURACIÓN INICIAL
# ---------------------------------------------------------
EXCEL_PATH = r"C:\Users\trast\OneDrive - Trast\Rendiciones\Rendiciones.xlsx"
PDF_DIR = r"C:\Users\trast\OneDrive - Trast\Rendiciones\Comprobantes_pdf"
os.makedirs(PDF_DIR, exist_ok=True)

# ---------------------------------------------------------
# CREAR EXCEL SI NO EXISTE
# ---------------------------------------------------------
if not os.path.exists(EXCEL_PATH):
    df_init = pd.DataFrame(columns=["Conductor", "Fecha", "Monto", "Descripcion", "LinkComprobante"])
    df_init.to_excel(EXCEL_PATH, index=False)

# ---------------------------------------------------------
# AJUSTAR COLUMNAS EXCEL
# ---------------------------------------------------------
def ajustar_columnas_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(path)

# ---------------------------------------------------------
# GUARDAR EN EXCEL
# ---------------------------------------------------------
def guardar_en_excel(conductor, fecha, monto, descripcion, link_pdf):
    df = pd.read_excel(EXCEL_PATH)

    nuevo = pd.DataFrame([{
        "Conductor": conductor,
        "Fecha": fecha,
        "Monto": monto,
        "Descripcion": descripcion,
        "LinkComprobante": link_pdf
    }])

    df = pd.concat([df, nuevo], ignore_index=True)
    df.to_excel(EXCEL_PATH, index=False)

    ajustar_columnas_excel(EXCEL_PATH)

# ---------------------------------------------------------
# GENERAR PDF (VERSIÓN MEJORADA)
# ---------------------------------------------------------
def generar_pdf(conductor, fecha, monto, descripcion, imagen_bytes):

    filename = f"{conductor}_{fecha.replace('/', '-')}.pdf"
    ruta_pdf = os.path.join(PDF_DIR, filename)

    doc = SimpleDocTemplate(ruta_pdf, pagesize=letter)
    styles = getSampleStyleSheet()

    story = []

    # Título
    titulo = "<para align='center'><b><font size=16>Control de Gastos Trast</font></b></para>"
    story.append(Paragraph(titulo, styles["Title"]))
    story.append(Spacer(1, 20))

    # Datos
    story.append(Paragraph(f"<b>Conductor:</b> {conductor}", styles["Normal"]))
    story.append(Paragraph(f"<b>Fecha:</b> {fecha}", styles["Normal"]))
    story.append(Paragraph(f"<b>Monto:</b> ${monto:,} CLP", styles["Normal"]))
    story.append(Paragraph(f"<b>Descripción:</b> {descripcion}", styles["Normal"]))
    story.append(Spacer(1, 20))

    # Guardar imagen temporal
    temp_path = os.path.join(PDF_DIR, "temp_image.jpg")
    with open(temp_path, "wb") as f:
        f.write(imagen_bytes)

    # Redimensionamiento automático
    img = RLImage(temp_path)

    MAX_WIDTH = 5.5 * inch
    MAX_HEIGHT = 6.5 * inch

    w_ratio = MAX_WIDTH / img.drawWidth
    h_ratio = MAX_HEIGHT / img.drawHeight
    scale = min(w_ratio, h_ratio, 1)

    img.drawWidth = img.drawWidth * scale
    img.drawHeight = img.drawHeight * scale

    story.append(img)

    # Crear PDF
    doc.build(story)

    # Borrar temp seguro
    try:
        os.remove(temp_path)
    except PermissionError:
        import time
        time.sleep(0.2)
        try:
            os.remove(temp_path)
        except:
            pass

    return ruta_pdf

# ---------------------------------------------------------
# INTERFAZ STREAMLIT
# ---------------------------------------------------------
st.title("📄 Sistema de Rendiciones - Trast")

st.write("Completa la información y adjunta una fotografía o imagen del comprobante.")

conductor = st.text_input("👤 Conductor")
fecha = st.date_input("📅 Fecha").strftime("%d/%m/%Y")
monto = st.number_input("💰 Monto (CLP)", min_value=0, step=100)
descripcion = st.text_area("📝 Descripción del gasto")

# Subir o sacar fotografía
uploaded_file = st.file_uploader("📸 Subir imagen del comprobante", type=["jpg", "jpeg", "png"])
camera_photo = st.camera_input("📷 Tomar fotografía del comprobante")

# ---------------------------------------------------------
# PROCESAR ENVÍO
# ---------------------------------------------------------
if st.button("Enviar Rendición", use_container_width=True):

    if not conductor or monto <= 0 or not descripcion:
        st.error("⚠ Por favor completa todos los campos.")
    else:
        imagen_bytes = None

        if uploaded_file:
            imagen_bytes = uploaded_file.read()
        elif camera_photo:
            imagen_bytes = camera_photo.getvalue()
        else:
            st.error("⚠ Debes subir o tomar una foto del comprobante.")
            st.stop()

        # Crear PDF
        ruta_pdf = generar_pdf(conductor, fecha, monto, descripcion, imagen_bytes)

        # Guardar en Excel
        guardar_en_excel(conductor, fecha, monto, descripcion, ruta_pdf)

        # Confirmación visual
        st.success("✅ La rendición fue enviada correctamente.")
        st.balloons()

        # Limpiar formulario
        st.rerun()
